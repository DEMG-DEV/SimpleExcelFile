from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from .DateDetails import *


class ExcelFile():

    def __init__(self):
        self.file = Workbook()
        self.sheet = self.file.active
        self.details = DateDetails()
        self.cells = ["A", "B", "C", "D", "E", "F"]
        self.acells = []
        self.work_hours = "9.5"
        self.date = None

    def set_datetime(self, datetime):
        self.date = datetime

    def save_as(self, filename):
        self.file.save(filename)

    def set_headers(self):
        self.sheet['A1'] = "Mes de " + self.date.strftime('%B')
        init_week, end_week = self.details.get_init_end_week(self.date.year, int(self.date.strftime('%V')), 4)

        end_month = self.details.get_month_day_range(self.date)

        if init_week < end_month.date() < end_week:
            self.sheet['A2'] = 'Week from ' + init_week.strftime('%d/%m/%y') + ' to ' + end_month.strftime(
                '%d/%m/%y')
        else:
            self.sheet['A2'] = 'Week from ' + init_week.strftime('%d/%m/%y') + ' to ' + end_week.strftime(
                '%d/%m/%y')

        self.sheet['A4'] = 'Activity'
        self.sheet['A5'] = 'Programación'
        for i in self.cells:
            if i != "A":
                self.sheet[i + '4'] = init_week.strftime('%A') + ' ' + str(init_week.day)
                try:
                    day = init_week.day + 1
                    init_week = date(init_week.year, init_week.month, day)
                    self.acells.append(i)
                except Exception as err:
                    self.acells.append(i)
                    self.cells.clear()
                    return

    def set_data(self):
        for i in self.acells:
            if i != "A":
                self.sheet[i + '5'] = self.work_hours
                self.sheet[i + '6'] = self.work_hours

    def get_week(self):
        return self.details.get_init_end_week(self.date.year, int(self.date.strftime('%V')), 4)

    def set_styles(self):
        a1 = self.sheet["A1"]
        a2 = self.sheet["A2"]
        ft = Font(name="Calibri", size="22", bold="True")
        a1.font = ft
        ft = Font(name="Calibri", size="16", bold="True")
        a2.font = ft
        ft = Font(name="Calibri", size="18", bold="True")
        fill = PatternFill(fill_type="solid", start_color='5b9bd5', end_color='5b9bd5')
        for i in self.acells:
            cell = self.sheet[i + '4']
            cell.font = ft
            cell.fill = fill
            if i != "A":
                self.sheet.column_dimensions[i].width = 21.11
            else:
                self.sheet.column_dimensions[i].width = 17.56
        ft = Font(name="Calibri", size="11")
        fill = PatternFill(fill_type="solid", start_color='ffc000', end_color='ffc000')
        for i in self.acells:
            cell = self.sheet[i + '5']
            cell.font = ft
            cell = self.sheet[i + '6']
            cell.font = ft
            if i != "A":
                cell.fill = fill
