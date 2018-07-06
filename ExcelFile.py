#!/usr/bin/env python
# _*_coding: utf-8_*_

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import datetime
from DateDetails import *


class ExcelFile():

    def __init__(self):
        self.file = Workbook()
        self.sheet = self.file.active
        self.details = DateDetails()
        self.cells = ["A", "B", "C", "D", "E", "F"]
        self.work_hours = "9.5"

    def set_datetime(self, datetime):
        self.date = datetime

    def save_as(self, filename):
        self.file.save(filename)

    def set_headers(self):
        self.sheet['A1'] = "Mes de " + self.date.strftime('%B')
        init_week, end_week = self.details.get_init_end_week(self.date.year, int(self.date.strftime('%V')), 4)
        self.sheet['A2'] = 'Week from ' + init_week.strftime('%d/%m/%y') + ' to ' + end_week.strftime('%d/%m/%y')
        self.sheet['A4'] = 'Activity'
        self.sheet['A5'] = 'Programación'
        for i in self.cells:
            if i != "A":
                self.sheet[i + '4'] = init_week.strftime('%A') + ' ' + str(init_week.day)
                day = init_week.day + 1
                init_week = date(init_week.year, init_week.month, day)

    def set_data(self):
        for i in self.cells:
            if i != "A":
                self.sheet[i + '5'] = self.work_hours
                self.sheet[i + '6'] = self.work_hours

    def get_week(self):
        return self.details.get_init_end_week(self.date.year, int(self.date.strftime('%V')), 4)

    def set_styles(self):
        a1 = self.sheet["A1"]
        a2 = self.sheet["A2"]
        ft = Font(name="Calibri", size="22", bold="True")
        a1.font = ft
        ft = Font(name="Calibri", size="16", bold="True")
        a2.font = ft
        ft = Font(name="Calibri", size="18", bold="True")
        fill = PatternFill(fill_type="solid", start_color='5b9bd5', end_color='5b9bd5')
        for i in self.cells:
            cell = self.sheet[i + '4']
            cell.font = ft
            cell.fill = fill
            if i != "A":
                self.sheet.column_dimensions[i].width = 21.11
            else:
                self.sheet.column_dimensions[i].width = 17.56
        ft = Font(name="Calibri", size="11")
        fill = PatternFill(fill_type="solid", start_color='ffc000', end_color='ffc000')
        for i in self.cells:
            cell = self.sheet[i + '5']
            cell.font = ft
            cell = self.sheet[i + '6']
            cell.font = ft
            if i != "A":
                cell.fill = fill
